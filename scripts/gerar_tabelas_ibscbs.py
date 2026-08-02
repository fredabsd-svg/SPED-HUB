#!/usr/bin/env python3
"""Gera `src/documentos/tabelas_ibscbs.json` das planilhas oficiais da SVRS.

A tabela de CST e de classificação tributária do IBS/CBS é publicada pela
SVRS em planilha, e muda: a versão 1.10 do IT 2025.002 incluiu seis códigos,
dividiu o 620004 em dois e renumerou o antigo 620005. Digitar isso à mão no
código seria errar no primeiro ato normativo — e errar em silêncio, porque
ninguém confere uma tabela de 164 linhas lendo.

Por isso a planilha oficial fica versionada em `dados/oficiais/` e o que o
programa lê é derivado dela por este script (§1.9). Atualizar a tabela é
baixar a planilha nova, trocar o arquivo e rodar isto — não editar Python.

    python scripts/gerar_tabelas_ibscbs.py

`tests/test_tabelas_ibscbs.py` refaz a geração e compara com o arquivo
versionado: planilha trocada sem regerar derruba o CI.
"""

from __future__ import annotations

import datetime as dt
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
DADOS = REPO / "dados" / "oficiais"
DESTINO = REPO / "src" / "documentos" / "tabelas_ibscbs.json"

ORIGEM = "https://dfe-portal.svrs.rs.gov.br/Nfe/Documentos"
DOCUMENTO = "IT 2025.002 v1.50 — Tabelas de Classificação do IBS e da CBS"

# As colunas são lidas pelo **nome do cabeçalho**, não pela posição: a
# planilha tem 82 colunas, dezenas delas vazias, e uma coluna inserida no
# meio deslocaria tudo sem erro nenhum.
INDICADORES_DO_CST = {
    "gIBSCBS": "ind_gIBSCBS",
    "gIBSCBSMono": "ind_gIBSCBSMono",
    "gRed": "ind_gRed",
    "gDif": "ind_gDif",
    "gTransfCred": "ind_gTransfCred",
    "gCredPresIBSZFM": "ind_ gCredPresIBSZFM",  # o espaço está na planilha
    "gAjusteCompet": "ind_gAjusteCompet",
    "RedutorBC": "ind_RedutorBC",
}


def _texto(valor: Any) -> str:
    return "" if valor is None else str(valor).strip()


def _bandeira(valor: Any) -> bool:
    return _texto(valor) == "1"


def _numero(valor: Any) -> float:
    texto = _texto(valor).replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def _data(valor: Any) -> str | None:
    if isinstance(valor, dt.datetime):
        return valor.date().isoformat()
    if isinstance(valor, dt.date):
        return valor.isoformat()
    texto = _texto(valor)
    return texto[:10] if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", texto) else None


def _aba(caminho: Path, prefixo: str):
    """A aba cujo título começa com `prefixo`.

    Os títulos carregam a data de publicação (`cClass 2026-06-01 Pub`), então
    procurar pelo nome exato quebraria na próxima planilha — que é justamente
    quando este script precisa funcionar.
    """
    wb = load_workbook(caminho, read_only=True, data_only=True)
    for ws in wb.worksheets:
        if ws.title.startswith(prefixo):
            linhas = [linha for linha in ws.iter_rows(values_only=True)]
            wb.close()
            cabecalho = [_texto(c) for c in linhas[0]]
            return cabecalho, linhas[1:]
    wb.close()
    raise SystemExit(f"{caminho.name}: nenhuma aba começando com {prefixo!r}")


def _coluna(cabecalho: list[str], nome: str) -> int:
    try:
        return cabecalho.index(nome)
    except ValueError:
        raise SystemExit(
            f"coluna {nome!r} ausente da planilha; cabeçalho: {[c for c in cabecalho if c]}"
        ) from None


def tabela_cst(caminho: Path) -> dict[str, dict]:
    cabecalho, linhas = _aba(caminho, "CST ")
    codigo = _coluna(cabecalho, "CST-IBS/CBS")
    descricao = _coluna(cabecalho, "Descrição CST-IBS/CBS")
    indices = {g: _coluna(cabecalho, c) for g, c in INDICADORES_DO_CST.items()}
    tabela: dict[str, dict] = {}
    for linha in linhas:
        cst = _texto(linha[codigo])
        if not re.fullmatch(r"\d{3}", cst):
            continue
        tabela[cst] = {
            "descricao": _texto(linha[descricao]),
            "exige": {grupo: _bandeira(linha[i]) for grupo, i in sorted(indices.items())},
        }
    return dict(sorted(tabela.items()))


def tabela_class_trib(caminho: Path) -> dict[str, dict]:
    cabecalho, linhas = _aba(caminho, "cClass")
    col = {
        nome: _coluna(cabecalho, nome)
        for nome in (
            "CST-IBS/CBS",
            "cClassTrib",
            "Nome cClassTrib",
            "Tipo de Alíquota",
            "pRedIBS",
            "pRedCBS",
            "dIniVig",
            "dFimVig",
            "indNFe",
            "indNFCe",
            "indNFSe",
        )
    }
    tabela: dict[str, dict] = {}
    for linha in linhas:
        codigo = _texto(linha[col["cClassTrib"]])
        if not re.fullmatch(r"\d{6}", codigo):
            continue
        tabela[codigo] = {
            "cst": _texto(linha[col["CST-IBS/CBS"]]),
            "nome": _texto(linha[col["Nome cClassTrib"]]),
            "tipo_aliquota": _texto(linha[col["Tipo de Alíquota"]]),
            "reducao_ibs": _numero(linha[col["pRedIBS"]]),
            "reducao_cbs": _numero(linha[col["pRedCBS"]]),
            "inicio_vigencia": _data(linha[col["dIniVig"]]),
            "fim_vigencia": _data(linha[col["dFimVig"]]),
            "na_nfe": _bandeira(linha[col["indNFe"]]),
            "na_nfce": _bandeira(linha[col["indNFCe"]]),
            "na_nfse": _bandeira(linha[col["indNFSe"]]),
        }
    return dict(sorted(tabela.items()))


def tabela_cred_pres(caminho: Path) -> dict[str, dict]:
    cabecalho, linhas = _aba(caminho, "cCredPres")
    col = {
        nome: _coluna(cabecalho, nome)
        for nome in (
            "cCredPres",
            "Descrição",
            "LC 214/2025",
            "Apropria via NF?",
            "Apropria via evento?",
        )
    }
    tabela: dict[str, dict] = {}
    for linha in linhas:
        codigo = _texto(linha[col["cCredPres"]])
        if not re.fullmatch(r"\d+", codigo):
            continue
        tabela[codigo.zfill(2)] = {
            "descricao": _texto(linha[col["Descrição"]]),
            "dispositivo": _texto(linha[col["LC 214/2025"]])[:120],
            "via_documento": _bandeira(linha[col["Apropria via NF?"]]),
            "via_evento": _bandeira(linha[col["Apropria via evento?"]]),
        }
    return dict(sorted(tabela.items()))


def _publicado_em(caminho: Path) -> str:
    """A data vem do **nome do arquivo oficial**, não do sistema de arquivos.

    `mtime` seria a data do download, e um `git clone` a reescreve — o
    programa passaria a dizer que a tabela é de hoje.
    """
    achado = re.search(r"(\d{4}-\d{2}-\d{2})", caminho.name)
    if not achado:
        raise SystemExit(f"{caminho.name}: nome sem a data de publicação (AAAA-MM-DD)")
    return achado.group(1)


def gerar() -> dict:
    classificacao = DADOS / "cClassTrib_2026-06-22.xlsx"
    credito = DADOS / "cCredPres_2026-06-22.xlsx"
    for arquivo in (classificacao, credito):
        if not arquivo.is_file():
            raise SystemExit(f"planilha oficial ausente: {arquivo.relative_to(REPO)}")
    return {
        "gerado_por": "scripts/gerar_tabelas_ibscbs.py",
        "aviso": "ARQUIVO GERADO — não edite; fonte: scripts/gerar_tabelas_ibscbs.py",
        "documento": DOCUMENTO,
        "origem": ORIGEM,
        "fontes": [
            {
                "arquivo": str(arquivo.relative_to(REPO)).replace("\\", "/"),
                "publicado_em": _publicado_em(arquivo),
            }
            for arquivo in (classificacao, credito)
        ],
        "cst": tabela_cst(classificacao),
        "class_trib": tabela_class_trib(classificacao),
        "cred_pres": tabela_cred_pres(credito),
    }


def serializar(tabelas: dict) -> str:
    return json.dumps(tabelas, ensure_ascii=False, indent=1, sort_keys=False) + "\n"


def main() -> int:
    tabelas = gerar()
    DESTINO.write_text(serializar(tabelas), "utf-8")
    print(
        f"{DESTINO.relative_to(REPO)}: "
        f"{len(tabelas['cst'])} CST, "
        f"{len(tabelas['class_trib'])} cClassTrib, "
        f"{len(tabelas['cred_pres'])} cCredPres"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
