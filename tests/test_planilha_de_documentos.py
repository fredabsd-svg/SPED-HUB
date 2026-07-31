"""Excel de mão dupla: exportar os itens, corrigir fora, reimportar.

A alteração em massa resolve "todos os itens com NCM 2203 viram CFOP 2102".
Não resolve o caso mais comum do saneamento: cada linha tem um valor diferente
e quem sabe qual é uma pessoa olhando.

O que estes testes protegem:

  * **a volta não grava.** `reimportar` devolve a mesma `Simulacao` de
    `simular`, e quem confirma é `confirmar`, num lote reversível. Uma
    planilha que gravasse ao ser lida seria a única escrita do sistema sem
    ninguém ver o que muda — e a que mais tem como dar errado, porque passou
    por um programa que não é este;
  * **a identidade é conferida, não presumida.** Planilha reordenada, com
    linha apagada ou vinda de outra base é o caso normal; casar por posição
    faria a correção cair no documento errado;
  * **só o que é editável volta**, e o resto é contexto para quem lê;
  * **o Excel mexe nos tipos.** `2102` volta como `2102.0`, e `1000,00` volta
    como `1000.0000000001`; os dois produziriam alteração onde não houve.
"""

from __future__ import annotations

import datetime
import io

import pytest
from openpyxl import Workbook, load_workbook

from src.db.models import (
    AjusteFiscal,
    DocumentoFiscal,
    Empresa,
    Escritorio,
    criar_engine,
    get_session,
    init_db,
)
from src.documentos import (
    COLUNAS,
    EDITAVEIS,
    ORIGEM_USUARIO,
    ImportadorDeDocumentos,
    PlanilhaInvalida,
    Selecao,
    aplicar_ajuste,
    confirmar,
    exportar,
    reimportar,
)
from tests.fixtures_nfe import nfe_xml


@pytest.fixture
def sessao(tmp_path):
    engine = criar_engine(url=f"sqlite:///{tmp_path / 'planilha.db'}")
    init_db(engine)
    with get_session(engine) as s:
        yield s


@pytest.fixture
def escritorio(sessao):
    e = Escritorio(nome="Teste", slug="teste")
    sessao.add(e)
    sessao.commit()
    return e


@pytest.fixture
def empresa(sessao, escritorio):
    e = Empresa(
        cnpj="98765432000198",
        nome="COMERCIO EXEMPLO LTDA",
        uf="TO",
        ie="293456789",
        cod_mun="1721000",
        escritorio_id=escritorio.id,
    )
    sessao.add(e)
    sessao.commit()
    return e


@pytest.fixture
def com_documento(sessao, escritorio, empresa):
    ocorrencia = ImportadorDeDocumentos(sessao, escritorio_id=escritorio.id).importar(
        nfe_xml(itens=2)
    )
    sessao.commit()
    return sessao.get(DocumentoFiscal, ocorrencia.documento_id)


def recorte(escritorio, empresa):
    return Selecao(escritorio_id=escritorio.id, empresa_id=empresa.id)


def ler(conteudo: bytes):
    """A planilha como uma lista de dicionários, do jeito que se lê."""
    aba = load_workbook(io.BytesIO(conteudo))["itens"]
    linhas = list(aba.iter_rows(values_only=True))
    return [dict(zip(linhas[0], linha, strict=True)) for linha in linhas[1:]]


def editar(conteudo: bytes, coluna: str, valor, linha: int = 2) -> bytes:
    """Muda uma célula, como quem edita no Excel.

    A atribuição é em `.value` e não pelo parâmetro de `cell()`: passar
    `value=None` ali é um no-op no openpyxl, e o teste de "apaguei a célula"
    passaria sem apagar nada.
    """
    livro = load_workbook(io.BytesIO(conteudo))
    aba = livro["itens"]
    cabecalho = [c.value for c in aba[1]]
    aba.cell(row=linha, column=cabecalho.index(coluna) + 1).value = valor
    buffer = io.BytesIO()
    livro.save(buffer)
    return buffer.getvalue()


# ── A ida ──────────────────────────────────────────────────────────────────


def test_exporta_uma_linha_por_item(sessao, escritorio, empresa, com_documento):
    linhas = ler(exportar(sessao, recorte(escritorio, empresa)))

    assert len(linhas) == 2
    assert [ln["numero_item"] for ln in linhas] == [1, 2]


def test_a_planilha_leva_a_identidade_e_o_contexto(sessao, escritorio, empresa, com_documento):
    """Sem contexto a planilha é uma tabela de ids, e ninguém a edita."""
    linha = ler(exportar(sessao, recorte(escritorio, empresa)))[0]

    assert linha["item_id"] == com_documento.itens[0].id
    assert linha["chave"] == com_documento.chave
    assert linha["descricao"] == "PRODUTO DE TESTE 1"


def test_exporta_o_efetivo_e_nao_o_normalizado(sessao, escritorio, empresa, com_documento):
    """Exportar o XML original desfaria as correções já feitas.

    Quem abre a planilha precisa ver o que iria para o SPED hoje.
    """
    aplicar_ajuste(
        sessao,
        documento=com_documento,
        item=com_documento.itens[0],
        campo="cfop",
        valor_novo="2102",
        origem=ORIGEM_USUARIO,
    )
    sessao.commit()

    linha = ler(exportar(sessao, recorte(escritorio, empresa)))[0]

    assert linha["cfop"] == "2102", "o XML trazia 6102"


def test_a_planilha_traz_as_colunas_declaradas(sessao, escritorio, empresa, com_documento):
    aba = load_workbook(io.BytesIO(exportar(sessao, recorte(escritorio, empresa))))["itens"]

    assert [c.value for c in aba[1]] == list(COLUNAS)


# ── A volta não grava ──────────────────────────────────────────────────────


def test_reimportar_nao_grava_nada(sessao, escritorio, empresa, com_documento):
    """A propriedade central: ler a planilha é propor, não gravar."""
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "cfop", "2102")

    reimportar(sessao, conteudo)
    sessao.commit()

    assert sessao.query(AjusteFiscal).count() == 0


def test_a_volta_devolve_a_simulacao_que_confirmar_grava(
    sessao, escritorio, empresa, com_documento
):
    """De ponta a ponta: a planilha é outra forma de propor a mesma coisa."""
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "cfop", "2102")

    resultado = reimportar(sessao, conteudo)
    lote = confirmar(sessao, resultado.simulacao, motivo="planilha de julho")
    sessao.commit()
    sessao.expire_all()

    assert lote
    item = sessao.get(DocumentoFiscal, com_documento.id).itens[0]
    ajuste = sessao.query(AjusteFiscal).filter_by(item_id=item.id, campo="cfop").one()
    assert ajuste.valor_novo == "2102"


def test_so_a_celula_mexida_vira_mudanca(sessao, escritorio, empresa, com_documento):
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "cfop", "2102")

    resultado = reimportar(sessao, conteudo)

    assert resultado.total_mudancas == 1
    assert resultado.linhas_lidas == 2


def test_planilha_intocada_nao_propoe_nada(sessao, escritorio, empresa, com_documento):
    """Se a ida e a volta não casassem, toda linha viraria alteração."""
    resultado = reimportar(sessao, exportar(sessao, recorte(escritorio, empresa)))

    assert resultado.total_mudancas == 0
    assert resultado.divergencias == []


def test_a_mudanca_sai_com_o_valor_anterior(sessao, escritorio, empresa, com_documento):
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "cfop", "2102")

    mudanca = reimportar(sessao, conteudo).simulacao.mudancas[0]

    assert mudanca.valor_anterior == "6102"
    assert mudanca.valor_novo == "2102"
    assert mudanca.numero_item == 1


def test_a_volta_parte_do_efetivo(sessao, escritorio, empresa, com_documento):
    """Comparar com o normalizado reproporia a correção que já existe."""
    aplicar_ajuste(
        sessao,
        documento=com_documento,
        item=com_documento.itens[0],
        campo="cfop",
        valor_novo="2102",
        origem=ORIGEM_USUARIO,
    )
    sessao.commit()

    resultado = reimportar(sessao, exportar(sessao, recorte(escritorio, empresa)))

    assert resultado.total_mudancas == 0


# ── A identidade é conferida ───────────────────────────────────────────────


def test_linha_de_outra_base_e_recusada_com_a_razao(sessao, escritorio, empresa, com_documento):
    """A chave é reconferida: o id sozinho é um número que se digita por cima."""
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "chave", "9" * 44)

    resultado = reimportar(sessao, conteudo)

    assert resultado.total_mudancas == 0
    assert "outra base" in str(resultado.divergencias[0])


def test_item_inexistente_e_recusado_sem_derrubar_o_resto(
    sessao, escritorio, empresa, com_documento
):
    """Uma linha ruim não pode custar as outras novecentas."""
    conteudo = exportar(sessao, recorte(escritorio, empresa))
    conteudo = editar(conteudo, "item_id", 999999)
    conteudo = editar(conteudo, "cfop", "2102", linha=3)

    resultado = reimportar(sessao, conteudo)

    assert resultado.total_mudancas == 1, "a segunda linha passou"
    assert "não existe item #999999" in str(resultado.divergencias[0])


def test_item_id_vazio_e_recusado(sessao, escritorio, empresa, com_documento):
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "item_id", None)

    resultado = reimportar(sessao, conteudo)

    assert "item_id vazio" in str(resultado.divergencias[0])


def test_a_ordem_das_colunas_nao_importa(sessao, escritorio, empresa, com_documento):
    """Quem edita reordena; recusar por isso seria recusar o uso normal."""
    original = ler(exportar(sessao, recorte(escritorio, empresa)))
    livro = Workbook()
    aba = livro.active
    aba.title = "itens"
    invertidas = list(reversed(COLUNAS))
    aba.append(invertidas)
    for linha in original:
        aba.append([linha[c] for c in invertidas])
    aba.cell(row=2, column=invertidas.index("cfop") + 1, value="2102")
    buffer = io.BytesIO()
    livro.save(buffer)

    resultado = reimportar(sessao, buffer.getvalue())

    assert resultado.total_mudancas == 1


def test_planilha_sem_a_identidade_e_recusada(sessao):
    livro = Workbook()
    aba = livro.active
    aba.title = "itens"
    aba.append(["chave", "cfop"])
    aba.append(["1" * 44, "2102"])
    buffer = io.BytesIO()
    livro.save(buffer)

    with pytest.raises(PlanilhaInvalida, match="item_id"):
        reimportar(sessao, buffer.getvalue())


def test_planilha_de_outra_aba_e_recusada(sessao):
    livro = Workbook()
    livro.active.title = "Planilha1"
    buffer = io.BytesIO()
    livro.save(buffer)

    with pytest.raises(PlanilhaInvalida, match="aba 'itens'"):
        reimportar(sessao, buffer.getvalue())


def test_arquivo_que_nao_e_planilha_e_recusado(sessao):
    with pytest.raises(PlanilhaInvalida, match="não foi possível abrir"):
        reimportar(sessao, b"isto nao e um xlsx")


# ── Só o que é editável volta ──────────────────────────────────────────────


def test_mudar_a_chave_na_planilha_nao_muda_a_chave(sessao, escritorio, empresa, com_documento):
    """Ela está lá para ser lida, não editada — e a volta diz isso."""
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "chave", "9" * 44)

    resultado = reimportar(sessao, conteudo)

    assert not [m for m in resultado.simulacao.mudancas if m.campo == "chave"]
    assert resultado.divergencias, "a linha foi recusada, não silenciosamente ignorada"


def test_mudar_a_descricao_nao_vira_alteracao(sessao, escritorio, empresa, com_documento):
    """A descrição é contexto: quem a corrige quer corrigir o cadastro."""
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "descricao", "OUTRO NOME")

    assert reimportar(sessao, conteudo).total_mudancas == 0


def test_coluna_apagada_da_planilha_nao_vira_alteracao(sessao, escritorio, empresa, com_documento):
    """Apagar a coluna é dizer "não mexi nisso", não "esvazie isso"."""
    original = ler(exportar(sessao, recorte(escritorio, empresa)))
    sem_ncm = [c for c in COLUNAS if c != "ncm"]
    livro = Workbook()
    aba = livro.active
    aba.title = "itens"
    aba.append(sem_ncm)
    for linha in original:
        aba.append([linha[c] for c in sem_ncm])
    buffer = io.BytesIO()
    livro.save(buffer)

    resultado = reimportar(sessao, buffer.getvalue())

    assert resultado.total_mudancas == 0


def test_todos_os_editaveis_sao_colunas_do_item(sessao):
    """Campo que não existe na tabela quebraria só na hora de reimportar."""
    from src.db.models import ItemDocumentoFiscal

    for campo in EDITAVEIS:
        assert campo in ItemDocumentoFiscal.__table__.columns, campo


# ── O Excel mexe nos tipos ─────────────────────────────────────────────────


def test_codigo_lido_como_numero_nao_vira_2102_ponto_zero(
    sessao, escritorio, empresa, com_documento
):
    """`2102.0` num CFOP produziria arquivo recusado pelo validador.

    O valor escrito é `float`, e não `int`: o openpyxl devolve inteiro como
    `int`, e a célula numérica do Excel volta como `float`. Testar com `int`
    não exercitaria a coerção nenhuma.
    """
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "cfop", 2102.0)

    mudanca = reimportar(sessao, conteudo).simulacao.mudancas[0]

    assert mudanca.valor_novo == "2102"


def test_diferenca_de_ponto_flutuante_nao_vira_alteracao(
    sessao, escritorio, empresa, com_documento
):
    """O Excel guarda 1000,00 como 1000.0000000001.

    Comparar por igualdade exata faria a planilha voltar propondo mudar tudo,
    e ninguém leria a lista.
    """
    conteudo = editar(
        exportar(sessao, recorte(escritorio, empresa)), "valor_total", 1000.0000000001
    )

    assert reimportar(sessao, conteudo).total_mudancas == 0


def test_meio_centavo_de_diferenca_ainda_e_alteracao(sessao, escritorio, empresa, com_documento):
    """A tolerância é do arredondamento, não uma licença para ignorar centavo."""
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "valor_total", 1000.01)

    assert reimportar(sessao, conteudo).total_mudancas == 1


def test_celula_esvaziada_apaga_o_campo(sessao, escritorio, empresa, com_documento):
    """Apagar o conteúdo da célula é diferente de apagar a coluna."""
    conteudo = editar(exportar(sessao, recorte(escritorio, empresa)), "cest", None)

    mudanca = reimportar(sessao, conteudo).simulacao.mudancas[0]

    assert mudanca.campo == "cest"
    assert mudanca.valor_novo is None


def test_data_que_o_excel_devolve_como_datetime_nao_quebra(
    sessao, escritorio, empresa, com_documento
):
    """Célula formatada como data volta `datetime`, não texto.

    Nenhum campo editável é coluna de data — formatar uma como data é engano
    de quem editou. O que importa é não quebrar: vira texto e segue.
    """
    conteudo = editar(
        exportar(sessao, recorte(escritorio, empresa)), "unidade", datetime.datetime(2026, 7, 1)
    )

    resultado = reimportar(sessao, conteudo)

    assert resultado.total_mudancas == 1
    assert resultado.divergencias == []


def test_linha_em_branco_no_fim_e_ignorada(sessao, escritorio, empresa, com_documento):
    """Excel sobra linha vazia ao fim de quase toda planilha editada.

    A linha é de strings vazias, não de `None`: o openpyxl não cria célula
    para `None`, e a linha simplesmente não existiria — o teste passaria sem
    exercitar nada.
    """
    livro = load_workbook(io.BytesIO(exportar(sessao, recorte(escritorio, empresa))))
    livro["itens"].append([""] * len(COLUNAS))
    buffer = io.BytesIO()
    livro.save(buffer)

    resultado = reimportar(sessao, buffer.getvalue())

    assert resultado.linhas_lidas == 2
    assert resultado.divergencias == []
