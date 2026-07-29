"""Testes da CLI (`sped-hub`).

`src/cli.py` são 386 statements e a primeira funcionalidade anunciada no
README — e estava com 0% de cobertura.  Os testes entram por ``main()`` com
``sys.argv`` trocado, exercitando também o parser de argumentos, e não só as
funções ``cmd_*``.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from src import cli
from src.db.models import criar_engine, get_session, init_db

FIXTURE = Path(__file__).parent / "fixtures" / "ecd_sample.txt"


def executar(*argv: str) -> None:
    """Roda a CLI como se tivesse sido chamada da linha de comando."""
    import sys

    original = sys.argv
    sys.argv = ["sped-hub", *argv]
    try:
        cli.main()
    finally:
        sys.argv = original


@pytest.fixture
def db(tmp_path) -> str:
    """Banco vazio, isolado por teste."""
    caminho = str(tmp_path / "cli.db")
    engine = criar_engine(caminho)
    try:
        init_db(engine)
    finally:
        engine.dispose()
    return caminho


@pytest.fixture
def db_com_ecd(db) -> str:
    """Banco com a ECD de fixture já importada."""
    executar("importar-ecd", str(FIXTURE), "--db", db)
    return db


class TestParseData:
    @pytest.mark.parametrize(
        "entrada,esperado",
        [("01012024", (2024, 1, 1)), ("2024-12-31", (2024, 12, 31))],
    )
    def test_formatos_aceitos(self, entrada, esperado):
        data = cli._parse_data(entrada)
        assert (data.year, data.month, data.day) == esperado

    def test_formato_invalido(self):
        with pytest.raises(ValueError):
            cli._parse_data("31/12/2024")


class TestImportarEcd:
    def test_importa_a_fixture(self, db):
        executar("importar-ecd", str(FIXTURE), "--db", db)

        engine = criar_engine(db)
        session = get_session(engine)
        try:
            from src.db.models import ECD, Lancamento, PlanoConta

            assert session.query(ECD).count() == 1
            assert session.query(PlanoConta).count() == 23
            assert session.query(Lancamento).count() == 9
        finally:
            session.close()
            engine.dispose()

    def test_arquivo_inexistente_sai_com_erro(self, db, tmp_path):
        with pytest.raises(SystemExit) as exc:
            executar("importar-ecd", str(tmp_path / "nao_existe.txt"), "--db", db)
        assert exc.value.code == 1

    def test_arquivo_invalido_sai_com_erro(self, db, tmp_path):
        lixo = tmp_path / "lixo.txt"
        lixo.write_text("isto não é uma ECD\n", encoding="utf-8")
        with pytest.raises(SystemExit) as exc:
            executar("importar-ecd", str(lixo), "--db", db)
        assert exc.value.code == 1

    def test_reimportar_e_recusado_sem_duplicar(self, db_com_ecd):
        """A segunda importação do mesmo arquivo é recusada, não duplicada.

        `ECDImportService` levanta `DuplicateECDImportError` e a CLI sai com 1.
        O que importa para a escrituração é o efeito: nada é duplicado.
        """
        with pytest.raises(SystemExit) as exc:
            executar("importar-ecd", str(FIXTURE), "--db", db_com_ecd)
        assert exc.value.code == 1

        engine = criar_engine(db_com_ecd)
        session = get_session(engine)
        try:
            from src.db.models import ECD, Lancamento, Partida

            assert session.query(ECD).count() == 1
            assert session.query(Lancamento).count() == 9
            assert session.query(Partida).count() == 18
        finally:
            session.close()
            engine.dispose()


class TestRelatorio:
    @pytest.mark.parametrize("tipo", ["balancete", "balanco", "dre", "diario"])
    def test_gera_sem_erro(self, db_com_ecd, capsys, tipo):
        executar("relatorio", tipo, "--db", db_com_ecd)
        assert capsys.readouterr().out.strip()

    def test_razao_exige_conta(self, db_com_ecd, capsys):
        executar("relatorio", "razao", "--conta", "1.1.1", "--db", db_com_ecd)
        assert "1.1.1" in capsys.readouterr().out

    def test_sem_ecd_importada_sai_com_erro(self, db):
        with pytest.raises(SystemExit) as exc:
            executar("relatorio", "balancete", "--db", db)
        assert exc.value.code == 1

    def test_balancete_confere(self, db_com_ecd, capsys):
        """O balancete precisa fechar — é a validação contábil mais básica."""
        executar("relatorio", "balancete", "--db", db_com_ecd)
        assert "Conferência: OK" in capsys.readouterr().out

    def test_filtro_por_natureza(self, db_com_ecd, capsys):
        executar("relatorio", "balancete", "--natureza", "01", "--db", db_com_ecd)
        saida = capsys.readouterr().out
        assert saida.strip()

    def test_ecd_id_inexistente(self, db_com_ecd):
        with pytest.raises(SystemExit) as exc:
            executar("relatorio", "balancete", "--ecd-id", "999", "--db", db_com_ecd)
        assert exc.value.code == 1


class TestExportar:
    @pytest.mark.parametrize("relatorio", ["balanco", "dre", "diario"])
    def test_pdf(self, db_com_ecd, tmp_path, relatorio):
        saida = tmp_path / f"{relatorio}.pdf"
        executar(
            "exportar",
            relatorio,
            "--formato",
            "pdf",
            "--saida",
            str(saida),
            "--db",
            db_com_ecd,
        )
        assert saida.exists()
        assert saida.read_bytes().startswith(b"%PDF"), "arquivo não é um PDF válido"

    def test_balancete_pdf_sai_no_caminho_pedido(self, db_com_ecd, tmp_path):
        """`--formato pdf` gera PDF de verdade, no caminho passado em `--saida`.

        Até a 0.16.x não existia template PDF para o balancete: o comando
        caía num XLSX gravado em `<saida>.xlsx`, avisando só em log INFO —
        quem automatizava recebia sucesso e não encontrava o arquivo. O
        template `balancete.html` fechou essa decisão de produto.
        """
        pedido = tmp_path / "balancete.pdf"
        executar(
            "exportar",
            "balancete",
            "--formato",
            "pdf",
            "--saida",
            str(pedido),
            "--db",
            db_com_ecd,
        )
        assert pedido.exists(), "o PDF não saiu no caminho pedido em --saida"
        assert pedido.read_bytes().startswith(b"%PDF"), "arquivo não é um PDF válido"
        antigo_substituto = tmp_path / "balancete.xlsx"
        assert (
            not antigo_substituto.exists()
        ), "o fallback antigo voltou: XLSX gravado em caminho diferente do pedido"

    @pytest.mark.parametrize("relatorio", ["balancete", "balanco", "dre", "diario"])
    def test_xlsx(self, db_com_ecd, tmp_path, relatorio):
        saida = tmp_path / f"{relatorio}.xlsx"
        executar(
            "exportar",
            relatorio,
            "--formato",
            "xlsx",
            "--saida",
            str(saida),
            "--db",
            db_com_ecd,
        )
        assert saida.exists()
        # XLSX é um zip: assinatura PK.
        assert saida.read_bytes().startswith(b"PK")

        from openpyxl import load_workbook

        wb = load_workbook(saida)
        assert wb.active.max_row > 1, "planilha sem linhas de dados"

    def test_white_label_aplica_escritorio(self, db_com_ecd, tmp_path):
        saida = tmp_path / "marca.pdf"
        executar(
            "exportar",
            "balanco",
            "--formato",
            "pdf",
            "--saida",
            str(saida),
            "--escritorio",
            "Escritorio Teste",
            "--cor",
            "#0B4F6C",
            "--db",
            db_com_ecd,
        )
        assert saida.exists() and saida.stat().st_size > 0


class TestValidar:
    def test_roda_e_reporta(self, db_com_ecd, capsys):
        executar("validar", "--db", db_com_ecd)
        saida = capsys.readouterr().out
        assert saida.strip()

    def test_sem_ecd_sai_com_erro(self, db):
        with pytest.raises(SystemExit) as exc:
            executar("validar", "--db", db)
        assert exc.value.code == 1


class TestFiltros:
    def test_listar_vazio(self, db_com_ecd, capsys):
        executar("filtros", "listar", "--db", db_com_ecd)
        assert "Nenhuma visão" in capsys.readouterr().out

    def test_salvar_listar_e_mostrar(self, db_com_ecd, capsys):
        criterios = json.dumps({"cod_nat": ["01"]})
        executar(
            "filtros",
            "salvar",
            "--nome",
            "so-ativo",
            "--criterios",
            criterios,
            "--db",
            db_com_ecd,
        )
        capsys.readouterr()

        executar("filtros", "listar", "--db", db_com_ecd)
        assert "so-ativo" in capsys.readouterr().out

        executar("filtros", "mostrar", "--nome", "so-ativo", "--db", db_com_ecd)
        assert "01" in capsys.readouterr().out


class TestInfo:
    def test_banco_vazio(self, db, capsys):
        executar("info", "--db", db)
        saida = capsys.readouterr().out
        assert "Empresas: 0" in saida

    def test_com_ecd(self, db_com_ecd, capsys):
        executar("info", "--db", db_com_ecd)
        saida = capsys.readouterr().out
        assert "Empresas: 1" in saida
        assert "EMPRESA EXEMPLO LTDA" in saida


class TestParserPrincipal:
    def test_sem_comando_mostra_ajuda(self, capsys):
        executar()
        assert "usage" in capsys.readouterr().out.lower()

    def test_comando_desconhecido_sai_com_erro(self):
        with pytest.raises(SystemExit) as exc:
            executar("nao-existe")
        assert exc.value.code != 0
