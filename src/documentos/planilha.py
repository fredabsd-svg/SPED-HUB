"""Excel de mão dupla: exportar os itens, corrigir fora, reimportar.

A alteração em massa resolve "todos os itens com NCM 2203 viram CFOP 2102".
Não resolve o outro caso, que é o mais comum no saneamento: **cada linha tem um
valor diferente e quem sabe qual é uma pessoa olhando**. Fazer isso item a item
na tela é inviável num mês com mil notas; fazer numa planilha é o que o
escritório já faz hoje, à mão, sem nada que traga o resultado de volta.

Três decisões dão forma ao módulo:

**A volta não grava.** `reimportar` devolve uma `Simulacao`, exatamente como
`simular`. Quem confirma é `confirmar`, num lote reversível. Uma planilha que
gravasse ao ser lida seria a única operação do sistema que altera escrituração
sem que ninguém veja o que vai mudar — e é a que mais tem como dar errado,
porque passou por um programa que não é este.

**A identidade viaja e é conferida.** Cada linha leva `documento_id` e
`item_id`, e a volta confere a chave da nota e o número do item contra o banco.
Planilha reordenada, com linha apagada ou colada de outro mês é o caso normal,
não o excepcional — casar por posição faria a correção de um documento cair em
outro, e o erro só apareceria na intimação.

**Só as colunas editáveis voltam.** A chave de acesso, o número da nota e o
CNPJ do emitente vão na planilha porque sem eles ninguém sabe o que está
editando — mas mudá-los ali não muda nada, e a volta diz isso em vez de fingir
que aceitou. A camada original continua intocada: reimportar vira ajuste, como
tudo o mais.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from src.db.models import ItemDocumentoFiscal
from src.documentos.ajustes import desserializar, valor_efetivo
from src.documentos.massa import Mudanca, Selecao, Simulacao, _ajustes_de

# As colunas que identificam a linha.  Vão para a planilha e voltam dela, mas
# como conferência: são o que liga a linha ao banco.
IDENTIDADE = ("documento_id", "item_id")

# O que a pessoa precisa ver para saber o que está editando, e que **não** é
# editável.  Sem elas a planilha é uma tabela de ids; com elas editáveis,
# alguém corrigiria a chave da NF-e num campo que o sistema não deve aceitar.
CONTEXTO = (
    "chave",
    "numero",
    "data_emissao",
    "sentido",
    "emitente_cnpj",
    "numero_item",
    "codigo",
    "descricao",
)

# O que se corrige numa planilha.  A lista é explícita, e não "todas as colunas
# do item": exportar as 68 colunas produziria uma planilha que ninguém lê, e
# aceitar todas de volta transformaria erro de digitação em qualquer campo do
# leiaute.
EDITAVEIS = (
    "cfop",
    "ncm",
    "cest",
    "origem_mercadoria",
    "cst_icms",
    "csosn",
    "cst_ipi",
    "cst_pis",
    "cst_cofins",
    # Os dois códigos da Reforma entram aqui pelo mesmo motivo que os de cima:
    # é o que uma nota importada traz errado e o escritório precisa corrigir.
    # A partir de 03/08/2026 eles passam a existir em toda nota, e sem eles a
    # planilha e a tela de correção alcançariam só o regime que está saindo.
    # Entram agora, e não antes, porque só agora um valor inventado é recusado
    # — a tabela oficial é o que separa código real de dígito digitado.
    "cst_ibscbs",
    "class_trib_ibscbs",
    "unidade",
    "quantidade",
    "valor_unitario",
    "valor_total",
    "valor_desconto",
    "base_icms",
    "aliquota_icms",
    "valor_icms",
    "valor_ipi",
    "valor_pis",
    "valor_cofins",
)

COLUNAS = (*IDENTIDADE, *CONTEXTO, *EDITAVEIS)

ABA = "itens"


class PlanilhaInvalida(ValueError):
    """A planilha não é a que saiu daqui, ou perdeu o que a liga ao banco."""


@dataclass
class Divergencia:
    """Linha que não pôde virar alteração, e por quê."""

    linha: int
    motivo: str

    def __str__(self) -> str:
        return f"linha {self.linha}: {self.motivo}"


@dataclass
class Reimportacao:
    """O que a planilha propõe — para ser lido antes de confirmar."""

    simulacao: Simulacao = field(default_factory=Simulacao)
    divergencias: list[Divergencia] = field(default_factory=list)
    linhas_lidas: int = 0

    @property
    def total_mudancas(self) -> int:
        return self.simulacao.total_mudancas


def exportar(session: Session, selecao: Selecao) -> bytes:
    """A planilha dos itens do recorte, já com a camada efetiva aplicada.

    Exporta o **efetivo**, não o normalizado: quem abre a planilha precisa ver
    o que iria para o SPED hoje, senão desfaz sem querer as correções que já
    foram feitas.
    """
    documentos = selecao.documentos(session)
    ajustes = _ajustes_de(session, documentos)

    livro = Workbook()
    aba = livro.active
    aba.title = ABA
    aba.append(list(COLUNAS))

    for documento in documentos:
        do_cabecalho = [a for a in ajustes.get(documento.id, []) if a.item_id is None]
        for item in documento.itens:
            do_item = [a for a in ajustes.get(documento.id, []) if a.item_id == item.id]
            aba.append(_linha(documento, item, do_cabecalho, do_item))

    buffer = io.BytesIO()
    livro.save(buffer)
    return buffer.getvalue()


def _linha(documento, item, do_cabecalho: list, do_item: list) -> list[Any]:
    valores: list[Any] = [documento.id, item.id]
    for campo in CONTEXTO:
        alvo, seus_ajustes = (
            (item, do_item)
            if campo in ItemDocumentoFiscal.__table__.columns
            else (documento, do_cabecalho)
        )
        valores.append(valor_efetivo(alvo, campo, seus_ajustes))
    for campo in EDITAVEIS:
        valores.append(valor_efetivo(item, campo, do_item))
    return valores


def reimportar(session: Session, conteudo: bytes) -> Reimportacao:
    """Lê a planilha e devolve o que ela mudaria — **sem gravar nada**.

    O resultado é a mesma `Simulacao` de `simular`, e por isso `confirmar`
    grava-a do mesmo jeito, no mesmo lote reversível. Uma planilha não é um
    caminho paralelo de escrita: é outra forma de propor a mesma coisa.
    """
    aba = _abrir(conteudo)
    cabecalho = _cabecalho(aba)

    resultado = Reimportacao()
    for numero_linha, linha in enumerate(aba.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c is not None and c != "" for c in linha):
            continue  # linha em branco no fim da planilha é o normal
        resultado.linhas_lidas += 1
        _ler_linha(session, dict(zip(cabecalho, linha, strict=False)), numero_linha, resultado)
    return resultado


def _abrir(conteudo: bytes):
    try:
        livro = load_workbook(io.BytesIO(conteudo), data_only=True)
    except Exception as erro:  # noqa: BLE001 — openpyxl levanta de tudo
        raise PlanilhaInvalida(f"não foi possível abrir a planilha: {erro}") from erro
    if ABA not in livro.sheetnames:
        raise PlanilhaInvalida(
            f"a planilha não tem a aba {ABA!r} — use a que saiu de `exportar`, "
            f"não uma nova (abas encontradas: {livro.sheetnames})"
        )
    return livro[ABA]


def _cabecalho(aba) -> list[str]:
    """Os nomes das colunas, conferidos contra o que a identidade exige.

    A ordem das colunas **não** importa: quem edita reordena, e recusar por
    isso seria recusar o uso normal. O que importa é que a identidade esteja
    lá — sem ela não há como saber a que item a linha se refere.
    """
    primeira = next(aba.iter_rows(min_row=1, max_row=1, values_only=True), ())
    nomes = [str(c).strip() if c is not None else "" for c in primeira]

    faltando = [c for c in IDENTIDADE if c not in nomes]
    if faltando:
        raise PlanilhaInvalida(
            f"a planilha não tem a(s) coluna(s) {', '.join(faltando)} — são elas que "
            "ligam cada linha ao documento no banco, e sem elas não dá para saber o "
            "que a linha está corrigindo"
        )
    return nomes


def _ler_linha(session: Session, valores: dict, numero_linha: int, resultado: Reimportacao) -> None:
    item = _achar_item(session, valores, numero_linha, resultado)
    if item is None:
        return

    ajustes = [
        a
        for a in _ajustes_de(session, [item.documento]).get(item.documento_id, [])
        if a.item_id == item.id
    ]

    for campo in EDITAVEIS:
        if campo not in valores:
            continue  # coluna que a pessoa apagou da planilha: não é alteração
        novo = _tipado(campo, valores[campo])
        atual = valor_efetivo(item, campo, ajustes)
        if _igual(atual, novo):
            continue
        resultado.simulacao.mudancas.append(
            Mudanca(
                documento_id=item.documento_id,
                chave=item.documento.chave,
                item_id=item.id,
                numero_item=item.numero_item,
                campo=campo,
                valor_anterior=atual,
                valor_novo=novo,
            )
        )


def _achar_item(
    session: Session, valores: dict, numero_linha: int, resultado: Reimportacao
) -> ItemDocumentoFiscal | None:
    """O item que a linha diz ser — conferindo que é mesmo ele.

    A chave de acesso é reconferida contra o banco. É o que impede que uma
    planilha de outro mês, ou com os ids editados, escreva no documento errado:
    o id sozinho é um número que qualquer um digita por cima.
    """
    try:
        item_id = int(valores.get("item_id"))
    except (TypeError, ValueError):
        resultado.divergencias.append(
            Divergencia(numero_linha, "item_id vazio ou não numérico — linha ignorada")
        )
        return None

    item = session.get(ItemDocumentoFiscal, item_id)
    if item is None:
        resultado.divergencias.append(
            Divergencia(numero_linha, f"não existe item #{item_id} — linha ignorada")
        )
        return None

    chave = valores.get("chave")
    if chave and str(chave).strip() != item.documento.chave:
        resultado.divergencias.append(
            Divergencia(
                numero_linha,
                f"a chave da linha não é a do item #{item_id} no banco — a planilha "
                "veio de outra base ou os identificadores foram editados",
            )
        )
        return None
    return item


def _tipado(campo: str, bruto: Any) -> Any:
    """O valor da célula no tipo da coluna.

    A conversão passa pelo mesmo `desserializar` que os ajustes usam — o
    mesmo texto tem de virar o mesmo valor, venha da planilha ou da tela.

    **Não há coerção de `float` para `int` aqui**, e a ausência é deliberada.
    O risco óbvio seria um CFOP numérico voltar como `2102.0` e virar a string
    `"2102.0"`, que o validador recusa. Só que tanto o Excel quanto o openpyxl
    normalizam número inteiro para `int` na leitura: `2102.0` volta `2102`, e
    `str()` já dá `"2102"`. Um ramo de coerção seria código que nenhuma entrada
    alcança, e código morto com comentário é pior que código nenhum — parece
    cobrir algo. Quem garante isso é
    `test_codigo_lido_como_numero_nao_vira_2102_ponto_zero`.

    Pelo mesmo motivo não há tratamento de `datetime`: nenhum campo editável é
    coluna de data.
    """
    coluna = ItemDocumentoFiscal.__table__.columns[campo]
    if bruto is None or bruto == "":
        return None
    return desserializar(str(bruto), coluna)


def _igual(atual: Any, novo: Any) -> bool:
    """Compara respeitando o centavo.

    O Excel guarda `1000.0000000001` para um `1000,00` digitado, e comparar por
    igualdade exata faria toda linha intocada virar alteração — a planilha
    voltaria propondo mudar tudo, e ninguém leria a lista.
    """
    if atual is None and novo is None:
        return True
    if isinstance(atual, int | float) and isinstance(novo, int | float):
        return abs(float(atual) - float(novo)) < 0.005
    return _texto(atual) == _texto(novo)


def _texto(valor: Any) -> str:
    return "" if valor is None else str(valor).strip()
