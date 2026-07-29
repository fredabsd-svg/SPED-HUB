"""Motor de Exportação (F18) — PDF via WeasyPrint + XLSX via openpyxl.

Renderiza templates Jinja2 com contexto e gera arquivos profissionais.
Suporta white-label (logo, cor primária, nome do escritório).
"""

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from jinja2 import Environment, FileSystemLoader, select_autoescape

from src.reports.base import fmt_data, fmt_data_hora, fmt_moeda

logger = logging.getLogger("sped-hub.export")


@dataclass
class WhiteLabel:
    """Configuração de white-label."""

    escritorio_nome: str = "SPED-HUB"
    cor_primaria: str = "#0B4F6C"
    cor_primaria_clara: str = "#E8F3F7"
    logo_path: str | None = None
    logo_base64: str | None = None

    def to_dict(self) -> dict:
        return {
            "escritorio_nome": self.escritorio_nome,
            "cor_primaria": self.cor_primaria,
            "cor_primaria_clara": self.cor_primaria_clara,
            "logo_base64": self.logo_base64,
        }


class ExportEngine:
    """Renderiza templates Jinja2 e exporta para PDF/XLSX."""

    def __init__(self, templates_dir: str | None = None):
        if templates_dir is None:
            templates_dir = str(Path(__file__).resolve().parent / "templates")
        self.env = Environment(
            loader=FileSystemLoader(templates_dir),
            autoescape=select_autoescape(["html"]),
        )
        # Registra filtros customizados
        self.env.globals["fmt_moeda"] = fmt_moeda
        self.env.globals["fmt_data"] = fmt_data
        self.env.globals["fmt_data_hora"] = fmt_data_hora

    def render_html(
        self,
        template_name: str,
        ctx: Any,
        white_label: WhiteLabel | None = None,
        **kwargs,
    ) -> str:
        """Renderiza template Jinja2 para HTML."""
        template = self.env.get_template(template_name)
        wl_dict = white_label.to_dict() if white_label else None
        return template.render(
            ctx=ctx,
            white_label=wl_dict,
            **kwargs,
        )

    def export_pdf(
        self,
        template_name: str,
        output_path: str,
        ctx: Any,
        white_label: WhiteLabel | None = None,
        **kwargs,
    ) -> str:
        """Renderiza template e exporta para PDF via WeasyPrint.

        Returns:
            Caminho do arquivo gerado.
        """
        html = self.render_html(template_name, ctx, white_label, **kwargs)

        try:
            from weasyprint import HTML

            base_url = str(Path(__file__).resolve().parent / "templates")
            HTML(string=html, base_url=base_url).write_pdf(output_path)
            logger.info("PDF gerado: %s", output_path)
            return output_path
        except ImportError:
            logger.error("WeasyPrint não instalado")
            raise
        except Exception as e:
            logger.error("Erro ao gerar PDF: %s", e)
            raise

    def export_xlsx(
        self,
        output_path: str,
        ctx: Any,
        linhas: list[dict],
        colunas: list[str],
        titulo: str = "",
        white_label: WhiteLabel | None = None,
    ) -> str:
        """Exporta dados tabulares para XLSX formatado.

        Args:
            output_path: Caminho do arquivo .xlsx
            ctx: ReportContext
            linhas: Lista de dicionários com os dados
            colunas: Lista de nomes de colunas (chaves dos dicionários)
            titulo: Título da planilha
            white_label: Configuração white-label

        Returns:
            Caminho do arquivo gerado.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Alignment,
                Border,
                Font,
                PatternFill,
                Side,
            )
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = titulo[:31] if titulo else "Relatório"

            # ── Estilos ──
            cor_primaria = white_label.cor_primaria.replace("#", "") if white_label else "0B4F6C"

            header_fill = PatternFill(
                start_color=cor_primaria, end_color=cor_primaria, fill_type="solid"
            )
            header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            zebra_fill = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin", color="D4D4DA"),
                right=Side(style="thin", color="D4D4DA"),
                top=Side(style="thin", color="D4D4DA"),
                bottom=Side(style="thin", color="D4D4DA"),
            )
            data_font = Font(name="Calibri", size=10)
            moeda_format = '#.##0,00;(#.##0,00);"-"'

            # ── Cabeçalho do relatório ──
            row = 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(colunas))
            ws.cell(row=row, column=1, value=titulo).font = Font(
                name="Calibri", size=14, bold=True, color=cor_primaria
            )
            row += 1

            if ctx:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(colunas))
                ws.cell(
                    row=row,
                    column=1,
                    value=f"{ctx.empresa_nome} — CNPJ: {ctx.empresa_cnpj} — Período: {ctx.periodo_ref}",
                ).font = Font(name="Calibri", size=9, color="6B6B85")
                row += 1

            row += 1  # blank row

            # ── Cabeçalho da tabela ──
            for col_idx, col_name in enumerate(colunas, 1):
                cell = ws.cell(row=row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            row += 1

            # ── Dados ──
            for i, linha in enumerate(linhas):
                for col_idx, col_name in enumerate(colunas, 1):
                    valor = linha.get(col_name, "")
                    cell = ws.cell(row=row, column=col_idx, value=valor)
                    cell.font = data_font
                    cell.border = thin_border

                    # Alinhamento
                    if isinstance(valor, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                        if col_name.lower() in (
                            "saldo",
                            "débito",
                            "crédito",
                            "valor",
                            "saldo_atual",
                            "saldo_anterior",
                            "debitos",
                            "creditos",
                            "saldo_inicial",
                            "saldo_final",
                        ):
                            cell.number_format = moeda_format
                    else:
                        cell.alignment = Alignment(horizontal="left")

                # Zebra
                if i % 2 == 1:
                    for col_idx in range(1, len(colunas) + 1):
                        ws.cell(row=row, column=col_idx).fill = zebra_fill

                row += 1

            # ── Ajuste de largura ──
            for col_idx in range(1, len(colunas) + 1):
                max_width = len(str(colunas[col_idx - 1])) + 4
                for linha in linhas:
                    val = str(linha.get(colunas[col_idx - 1], ""))
                    max_width = max(max_width, len(val) + 2)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, 50)

            # ── Freeze panes ──
            ws.freeze_panes = ws.cell(row=row - len(linhas) - 1, column=1)

            wb.save(output_path)
            logger.info("XLSX gerado: %s", output_path)
            return output_path

        except ImportError:
            logger.error("openpyxl não instalado")
            raise
        except Exception as e:
            logger.error("Erro ao gerar XLSX: %s", e)
            raise

    def export_xlsx_to_buffer(
        self,
        buffer,
        ctx: Any,
        linhas: list[dict],
        colunas: list[str],
        titulo: str = "",
        white_label=None,
    ):
        """Exporta dados tabulares para XLSX em buffer (BytesIO)."""
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = titulo[:31] if titulo else "Relatorio"

        cor_primaria = white_label.cor_primaria.replace("#", "") if white_label else "0B4F6C"

        header_fill = PatternFill(
            start_color=cor_primaria, end_color=cor_primaria, fill_type="solid"
        )
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        zebra_fill = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D4D4DA"),
            right=Side(style="thin", color="D4D4DA"),
            top=Side(style="thin", color="D4D4DA"),
            bottom=Side(style="thin", color="D4D4DA"),
        )
        data_font = Font(name="Calibri", size=10)
        moeda_format = '#.##0,00;(#.##0,00);"-"'

        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(colunas))
        ws.cell(row=row, column=1, value=titulo).font = Font(
            name="Calibri", size=14, bold=True, color=cor_primaria
        )
        row += 1

        if ctx:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(colunas))
            ws.cell(
                row=row,
                column=1,
                value=f"{ctx.empresa_nome} — CNPJ: {ctx.empresa_cnpj} — Periodo: {ctx.periodo_ref}",
            ).font = Font(name="Calibri", size=9, color="6B6B85")
            row += 1

        row += 1

        for col_idx, col_name in enumerate(colunas, 1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        row += 1

        for i, linha in enumerate(linhas):
            for col_idx, col_name in enumerate(colunas, 1):
                valor = linha.get(col_name, "")
                cell = ws.cell(row=row, column=col_idx, value=valor)
                cell.font = data_font
                cell.border = thin_border
                if isinstance(valor, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    if col_name.lower() in (
                        "saldo",
                        "debito",
                        "credito",
                        "valor",
                        "saldo_atual",
                        "saldo_anterior",
                        "debitos",
                        "creditos",
                        "saldo_inicial",
                        "saldo_final",
                    ):
                        cell.number_format = moeda_format
                else:
                    cell.alignment = Alignment(horizontal="left")
            if i % 2 == 1:
                for col_idx in range(1, len(colunas) + 1):
                    ws.cell(row=row, column=col_idx).fill = zebra_fill
            row += 1

        for col_idx in range(1, len(colunas) + 1):
            max_width = len(str(colunas[col_idx - 1])) + 4
            for linha in linhas:
                val = str(linha.get(colunas[col_idx - 1], ""))
                max_width = max(max_width, len(val) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, 50)

        wb.save(buffer)
